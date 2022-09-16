import contextlib
import pathlib

import openpyxl
from celery import shared_task
from django.conf import settings
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, validator, ValidationError, constr

from .models import Company, Offer, OffersStore


class FileExchange:

    def __init__(self, file_name: str):
        self._file_name = file_name

    @property
    def file_path(self) -> pathlib.Path:
        return settings.OFFER_FILES_ROOT / self._file_name

    def write(self, file_bytes: bytes):
        with open(self.file_path, 'wb') as file:
            file.write(file_bytes)


class OfferParsedRow(BaseModel):
    name: constr(min_length=1, max_length=255)
    sku: constr(min_length=1, max_length=255)
    brand: constr(min_length=1, max_length=255) | None
    price: int | None
    store_ids: list[str] | None

    @validator('store_ids', pre=True, allow_reuse=True)
    def store_ids_str_to_list(cls, value: str | None):
        if value is None:
            return
        if isinstance(value, str):
            return value.split(',')
        raise ValidationError('Not str type')


def generate_offers_xlsx(file_path: pathlib.Path | str, company: Company) -> None:
    wb = Workbook(write_only=True)
    with contextlib.closing(wb):
        ws = wb.create_sheet('Товары')
        ws.append(('Название товара', 'SKU', 'Бренд', 'Цена', 'ID складов (можно перечислять несколько через запятую)'))
        for offer in company.offer_set.all():
            store_ids = ','.join(available_store.marketplace_store_id
                                 for available_store in offer.available_stores.all())
            ws.append((offer.name, offer.sku, offer.brand, offer.price, store_ids))
        wb.save(file_path)


def retrieve_offers_from_worksheet(worksheet: Worksheet) -> list[list]:
    result = []
    row_number = 1
    while True:
        row_number += 1
        rows = worksheet[f'A{row_number}': f'E{row_number}']
        if not rows:
            break
        result.append([row_number] + [cell.value for cell in rows[0]])
    return result


@shared_task
def parse_offers_xlsx(file_id: str, company_id: int) -> list[OfferParsedRow]:
    file_exchange = FileExchange(file_id)
    workbook = openpyxl.load_workbook(file_exchange.file_path, read_only=True)

    offers: list[OfferParsedRow] = []

    with contextlib.closing(workbook):
        try:
            worksheet: Worksheet = workbook['Товары']
        except KeyError:
            return []
        offer_rows = retrieve_offers_from_worksheet(worksheet)

    for row in offer_rows:
        _, name, sku, brand, price, store_ids = row
        try:
            offer = OfferParsedRow(name=name, sku=sku, brand=brand, price=price, store_ids=store_ids)
        except ValidationError as e:
            print(str(e))
        else:
            offers.append(offer)
    company = Company.objects.get(id=company_id)
    offer_stores = {store.marketplace_store_id: store for store in OffersStore.objects.filter(user=company.user)}

    Offer.objects.filter(company__id=company_id).delete()
    for offer_data in offers:
        available_stores = []
        for store_id in offer_data.store_ids:
            if store_id not in offer_stores:
                continue
            available_stores.append(offer_stores[store_id])
        offer = Offer.objects.create(company=company, sku=offer_data.sku, name=offer_data.name, brand=offer_data.brand, price=offer_data.price)
        print(offer)
        if available_stores:
            offer.available_stores.add(*available_stores)
